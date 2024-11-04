import glob
from io import BytesIO
import os
import subprocess
import sys
import time
import openpyxl
from datetime import datetime
import streamlit as st
if 'page_config_set' not in st.session_state:
    st.set_page_config(page_title="Podklady Rozvrhů", layout="wide")
    st.session_state['page_config_set'] = True
from streamlit_js_eval import streamlit_js_eval
import pandas as pd
from urllib.parse import urlparse, urlunparse

import sources.config as config
from sources.setup import setupDirectory
import sources.global_functions as global_functions
import sources.stahovani as stahovani
import sources.file_operations.tvorba_finalniho_vysledku as tfv
import sources.file_operations.krouzky as krouzky
import sources.file_operations.rozvrhove_akce as ra
import sources.file_operations.zakladni_tabulky as zt


def main():

    #Vytvoření podsložek
    setupDirectory()

    st.markdown("""
    <style>
    /* Style for the specific button with the custom ID */
    #enlarged-button > button {
        font-size: 20px;  /* Larger font */
        padding: 15px 30px;  /* Bigger padding for larger button */
        width: 200px;  /* Set a larger width */
    }
    </style>
    <style>
    [data-testid="stFileUploaderDropzone"] div div::before {content:"Zde přetáhni soubor"}
    [data-testid="stFileUploaderDropzone"] div div span{display:none;}
    [data-testid="stFileUploaderDropzone"] div div::after {font-size: .8em; content:"Limit 20MB"}
    [data-testid="stFileUploaderDropzone"] div div small{display:none;}
    div[data-testid="stFileUploader"]>section[data-testid="stFileUploaderDropzone"]>button[data-testid="baseButton-secondary"] {
    color:white;
    }
    div[data-testid="stFileUploader"]>section[data-testid="stFileUploaderDropzone"]>button[data-testid="baseButton-secondary"]::after {
    content: "Procházet...";
    color:black;
    display: block;
    position: absolute;
    }
    </style>
    """, unsafe_allow_html=True)
    #login_url = "https://stag-demo.zcu.cz/ws/login?originalURL=http://localhost:8501/"

    if "uploader_key" not in st.session_state:
        st.session_state.uploader_key = 0

    if not 'stagUserTicket' in st.query_params: #nepřihlášený uživatel
        login_url = "https://ws.ujep.cz/ws/login?originalURL=http://localhost:8501/" 
        st.markdown('# Přihlašte se pomocí <a href="https://ws.ujep.cz/ws/login?originalURL=http://localhost:8501/" target="_self">STAGu</a>', unsafe_allow_html=True)
    else:
        getKatedraList(2023)
        st.title('Podklady')
        st.text("")
        st.text("")
        st.text("")
        st.text("")
        widget_id = (id for id in range(1, 100_00))
        colA, colB, colC, colD, colE = st.columns([1,1,1,1,1])
        with colA:
            rok = st.selectbox(
            "Rok",
            options=(2023,2024), 
            key=next(widget_id))
        with colB:
            semestr = st.selectbox(
            "Semestr",
            ("LS", "ZS"), 
            key=next(widget_id))
        with colC:
            katedra = st.selectbox(
                "Katedra",
                getKatedraList(rok),
                key=next(widget_id))
        st.markdown('<div class="right-align">', unsafe_allow_html=True)
        with colE:
            st.markdown("""
                <div style="display: flex; flex-direction: column; align-items: center; justify-content: center;">
                """, unsafe_allow_html=True)
            with st.form("zaklady_form"): # Generace složeného souboru, ze kterého je pozdějí vytvářen hledaný výsledek
                if st.form_submit_button('Obnov seznam předmětů'):
                    with st.spinner("Probíhá příkaz..."):

                        files = glob.glob(config.folder_vysledky + '*')
                        for f in files:
                            os.remove(f)

                        start = time.time()
                        vsechny_fakulty=["FF","FSC","FSE","FSI","FUD","FZS","FŽP","PF","PFC","PRF","REK","RZS","U3V","UHS","UPV","UZM","UZS","UZT","UZU","ÚTŘ","IVK"]
                        stahovani.stahni_krouzky(rok)
                        for fakulta in vsechny_fakulty: 
                            print("FAKULTA: " + fakulta)  
                            stahovani.stahni_studijni_programy(fakulta, rok) 
                            programy=pd.read_csv(global_functions.getProgramySoubor(fakulta), sep=config.separator, engine='python', dtype=str)
                            for iterator1, line1 in programy.iterrows():
                                print(" --- Program " + str(iterator1+1) + "/" + str(len(programy)) + " --- ")
                                if(line1['typ']=='Doktorský'):
                                    print("Přeskočeno (Doktorský program).")
                                    continue
                                program_idno = global_functions.ziskej_program_data(fakulta, iterator1, "stprIdno")
                                program_kod = global_functions.ziskej_program_data(fakulta, iterator1, "kod")
                                stahovani.stahni_obory_programu(program_idno)
                                obory=pd.read_csv(global_functions.getOborySoubor(program_idno), sep=config.separator, engine='python', dtype=str)
                                for iterator2, line2 in obory.iterrows():
                                    obor_idno = global_functions.ziskej_obor_data(iterator2, program_idno , "oborIdno")
                                    obor_cislo = global_functions.ziskej_obor_data(iterator2, program_idno , "cisloOboru")
                                    stahovani.stahni_predmety_oboru(obor_idno, rok)
                                    zt.zkombinuj_do_vysledku(obor_idno, program_idno, fakulta, program_kod, obor_cislo, rok, semestr)
                        end = time.time()
                        print("Program bezel " + str(end - start) + " sekund.")
                        start = time.time()
                        zt.zkombinuj_vysledky(rok)
                        end = time.time()
                        print("Program bezel " + str(end - start) + " sekund.")
                file_path = global_functions.getSlozenyVysledek(rok)
                mod_time = os.path.getmtime(file_path)
                mod_time_formatted = datetime.fromtimestamp(mod_time).strftime('%d.%m.%Y')
                st.text("Poslední obnovení:" + mod_time_formatted)
            st.markdown("</div>", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)



        col0, col1, col2 = st.columns([5,0.5,0.5])
        #Stažení vzorového souboru pro doplnění vlastních kapacit
        #a souboru s kroužky
        st.markdown('<div class="right-align">', unsafe_allow_html=True)
        with col1:
            st.text("")
            st.text("")
            with open(config.dest_kapacity, 'rb') as file:
                file_bytes = file.read()
            st.download_button(
                label='Vlastní kapacita (k úpravě)',
                data=file_bytes,
                file_name='kapacity.xlsx',
                mime='application/vnd.ms-excel'
            )
        with col2:
            st.text("")
            st.text("")
            with open(config.dest_krouzky, 'rb') as file:
                file_bytes = file.read()
            st.download_button(
                label='Vlastní počty (k úpravě)',
                data=file_bytes,
                file_name='krouzky.xlsx',
                mime='application/vnd.ms-excel'
            )
        st.markdown('</div>', unsafe_allow_html=True)

        with st.form("my_form"): #nezávislé vstupy jsou uvnitř formuláře kvůli zabránění zbytečného obnovování stránky při změnách
            st.subheader('Další vstupy')
            col3, col4, col5, col6, col7, col8 = st.columns([0.3,0.3,0.3,0.7,0.7,0.7])
            with col3:
                config.prednaska_kapacita = st.number_input(label="Kapacita přednášek", value=9999, min_value=1) 
            with col4:
                config.cviceni_kapacita = st.number_input(label="Kapacita cvičení", value=30, min_value=1)
            with col5:
                config.seminar_kapacita = st.number_input(label="Kapacita seminářů", value=30, min_value=1)
            with col6:
                #Uživatel má možnost zadat vlastní kapacitu dle předmětu a dle typu akce.
                #Soubor je ve formátu XLSX a je při nahrávání validovaný.
                #Pokud validace projde, bere se v potaz ve funkci "rozdel_na_rozvrhove_akce" v souboru rozvrhove_akce.py
                vlastni_kapacita = st.file_uploader("Vlastní kapacita dle předmětu", type="xlsx")
                if vlastni_kapacita is not None:
                    df1=pd.read_excel(vlastni_kapacita, dtype=str)
                    #Validace
                    if(
                        len(df1.columns)!=4 or
                        df1.columns[0] != 'zkratka' or
                        df1.columns[1] != 'prednaska' or
                        df1.columns[2] != 'cviceni' or
                        df1.columns[3] != 'seminar' 
                    ):
                        st.warning('Soubor neprošel validací.')
                        st.session_state.uploader_key += 1
                        st.rerun()
                    else:
                        df1.to_excel(config.dest_kapacity, index=False)
            with col7:
                #Uživatel má možnost přistoupit k souboru s kroužky
                #a libovolně měnit hodnoty počtů studentů.
                vlastni_pocty = st.file_uploader("Vlastní počty studentů dle kroužku", type="xlsx")
                if vlastni_pocty is not None:
                    df2=pd.read_excel(vlastni_pocty, dtype=str)
                    if(
                        len(df2.columns)!=11 or
                        df2.columns[0] != 'Kód kroužku' or
                        df2.columns[1] != 'Popis' or
                        df2.columns[2] != 'Ročník' or
                        df2.columns[3] != 'Místo výuky' or
                        df2.columns[4] != 'Fakulta' or
                        df2.columns[5] != 'Program' or
                        df2.columns[6] != 'Obor' or
                        df2.columns[7] != 'Kombinace' or
                        df2.columns[8] != 'Rok' or
                        df2.columns[9] != 'Počet studentů kroužku' or
                        df2.columns[10] != 'Forma' 
                    ):
                        st.warning('Soubor neprošel validací.')
                        st.session_state.uploader_key += 1
                        st.rerun()
                    else:
                        df2.to_excel(config.dest_krouzky, index=False)
            
            with col8:
                #Hlavní část programu. Využívá soubor slozenyVysledek pro daný rok,
                #a společně se vstupním semestrem a katedrou postupně vygeneruje
                #podkladový soubor.
                #Graficky zobrazené v doprovodném diagramu "podklady_generaceDiagram".
                st.text("")
                st.text("")
                if st.form_submit_button('Generuj'):
                    with st.spinner("1/5 Hledání předmětů... "):
                        start = time.time()
                        print(type(rok))
                        print("katedra = " + katedra + ", Semestr = " + semestr + ", Rok = " + str(rok))
                        df = zt.vysledek_pro_katedru(katedra, semestr, rok)
                    with st.spinner("2/5 Stahování oborů předmětů... "):
                        stahovani.stahni_obory_predmetu(df, katedra, semestr, rok)
                    with st.spinner("3/5 Přidávání kroužků... "):
                        df = krouzky.krouzky_a_forma_z_oboru_predmetu(df,katedra, semestr, rok)
                    with st.spinner("4/5 Dělení na rozvrhové akce... "):
                        df = ra.rozdel_na_rozvrhove_akce(df, katedra, semestr, rok)
                    with st.spinner("5/5 Hledání spol. výuky a finální úpravy... "):
                        tfv.rozdel_vysledny_soubor(df, katedra, semestr, rok)
                        print("Konec")
                        end = time.time()
                        print("Program běžel " + str(end - start) + " sekund.")

        
        col1, col2 = st.columns([1,1])
        with col1: #pokud pro zadaný vstup existuje výsledek, zobrazí možnost jej stáhnout
            if(os.path.isfile(global_functions.getVysledekKatedry(katedra, semestr, rok))):
                st.text("")
                st.text("")
                with open(global_functions.getVysledekKatedry(katedra, semestr, rok), 'rb') as file:
                    file_bytes = file.read()
                st.download_button(
                    label='Stáhni výsledek XLSX',
                    data=file_bytes,
                    file_name='vysledek.xlsx',
                    mime='application/vnd.ms-excel'
                )
        
        # zobrazení výsledné tabulky na webu - děleno na záložky. 
        if(os.path.isfile(global_functions.getVysledekKatedry(katedra, semestr, rok))):
            tab_prez, tab_komb, tab_mix, tab_zatez = st.tabs(["Prezenční", "Kombinované", "Mix", "Zátěž"])
            try: #Pokud není nalezen list s prezenční výukou, nic se nezobrazí. TODO umožnit zobrazení pokud jsou jen kombinované nebo mix rozvrhové akce
                df = pd.read_excel(global_functions.getVysledekKatedry(katedra, semestr, rok), "prezencni", dtype=str)
                df = df.drop(['prednasejiciSPodily', 'cviciciSPodily'], axis=1)
                if('seminariciSPodily' in df):
                    df = df.drop(['seminariciSPodily'], axis=1)
            except:
                print("Prezencni list nenalezen.")
            else:
                #vzhledové úpravy
                global_column_config={
                    "jednotekPrednasek": st.column_config.Column(
                        "jednotkyPr",
                        help="Jednotky přednášek",
                        width="small"
                    ),
                    "jednotkaPrednasky": st.column_config.Column(
                        "jednotkaPr",
                        help="Jednotka přednášky",
                        width="small"
                    ),
                    "jednotekCviceni": st.column_config.Column(
                        "jednotekCv",
                        help="Jednotky cvičení",
                        width="small"
                    ),
                    "jednotkaCviceni": st.column_config.Column(
                        "jednotkaCv",
                        help="Jednotka cvičení",
                        width="small"
                    ),
                    "doporucenyRocnik": st.column_config.Column(
                        "Dop. ročník",
                        help="Doporučený ročník",
                        width="small"
                    ),
                    "krouzky": st.column_config.Column(
                        "kroužky",
                        help="Kroužky spojené s rozvrhovou akcí",
                        width="small"
                    ),
                    "spolecnaVyuka": st.column_config.Column(
                        "Společná výuka",
                        help="RA se stejným číslem značí, že jde o společnou výuku",
                        width="small"
                    ),
                    "zdrojVyucujiciho": st.column_config.Column(
                        "Zdroj vyučujícího",
                        help="Jak byl k akci zvolen vyučující",
                        width="small"
                    )

                }
                with tab_prez:
                    st.dataframe(df, column_config=global_column_config)
                with tab_komb:
                    df = pd.read_excel(global_functions.getVysledekKatedry(katedra, semestr, rok), "kombinovana", dtype=str)
                    df = df.drop(['prednasejiciSPodily', 'cviciciSPodily'], axis=1)
                    if('seminariciSPodily' in df):
                        df = df.drop(['seminariciSPodily'], axis=1)
                    st.dataframe(df, column_config=global_column_config)

                with tab_mix:
                    df = pd.read_excel(global_functions.getVysledekKatedry(katedra, semestr, rok), "mix", dtype=str)
                    df = df.drop(['prednasejiciSPodily', 'cviciciSPodily'], axis=1)
                    if('seminariciSPodily' in df):
                        df = df.drop(['seminariciSPodily'], axis=1)
                    st.dataframe(df, column_config=global_column_config)

                with tab_zatez:
                    df = pd.read_excel(global_functions.getVysledekKatedry(katedra, semestr, rok), "zatez", dtype=str)
                    st.dataframe(df)

def get_user_ticket():
    if not 'stagUserTicket' in st.query_params:
        return ''
    else:
        return st.query_params['stagUserTicket']

def refresh_url():
    st.query_params.clear()
    st.rerun()


    #Funkce vrací seznam všech kateder, které mají v daném roce alespoň jeden předmět
    #(a tudíž je má cenu nastavit jako možnost pro generování podkladů).
    #Jakmile takový seznam vytvoří, uloží ho jako nový list ke složenému výsledku,
    #aby se hledání nemuselo opakovat.
def getKatedraList(rok):
    book = openpyxl.load_workbook(global_functions.getSlozenyVysledek(rok))
    if not 'katedry' in book.sheetnames: 
        book.create_sheet('katedry')
        book.save(global_functions.getSlozenyVysledek(rok))
        book.close()
        df_temp_katedra = pd.read_excel(global_functions.getSlozenyVysledek(rok), sheet_name='Sheet1', dtype=str)
        df_temp_katedraList = df_temp_katedra.katedra.unique().tolist()
        df_temp_katedraList.sort()
        df = pd.DataFrame({'katedra':df_temp_katedraList})
        df1 = pd.read_excel(global_functions.getSlozenyVysledek(rok), dtype=str)
        writer = pd.ExcelWriter(global_functions.getSlozenyVysledek(rok)) #TODO: udělat celý proces pomocí už vytvořeného "book". Používání pandas writeru je potenciálně zbytečná komplikace
        df1.to_excel(writer, sheet_name='Sheet1')
        df.to_excel(writer,sheet_name='katedry')
        writer.close()
    else:
        book.close()
    df = pd.read_excel(global_functions.getSlozenyVysledek(rok), sheet_name='katedry', dtype=str)
    return df['katedra'].tolist()

# Zajišťuje, že se Streamlit zapne pouze jednou
# def run_streamlit():
#     if "streamlit" not in sys.argv[0]:
#         script_path = os.path.abspath(__file__)
#         subprocess.run([sys.executable, "-m", "streamlit", "run", script_path])

if __name__ == "__main__":
    main()
